#!/usr/bin/env python3
"""Validate the deterministic AITIP-0001 package and returned Claude workbook.

This script never calls Claude or any provider. A PASS is technical test evidence only;
it does not change the candidate/system/human decision state.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
FIXTURE_PATH = REPO_ROOT / "data" / "fixtures" / "AITIP-0001_sales.csv"
EXPECTED_PATH = REPO_ROOT / "data" / "fixtures" / "AITIP-0001_expected.json"
DEFAULT_WORKBOOK = (
    REPO_ROOT
    / "generated"
    / "candidates"
    / "AITIP-0001"
    / "AITIP-0001_claude-output.xlsx"
)


class ContractError(Exception):
    pass


def fail(message: str) -> None:
    raise ContractError(message)


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def number(value: Any, label: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        fail(f"{label} must be numeric; got {value!r}")
    return float(value)


def iso_date(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def load_package() -> tuple[list[dict[str, str]], dict[str, Any]]:
    if not FIXTURE_PATH.is_file():
        fail(f"Fixture missing: {FIXTURE_PATH}")
    if not EXPECTED_PATH.is_file():
        fail(f"Ground truth missing: {EXPECTED_PATH}")

    with FIXTURE_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        headers = reader.fieldnames or []
    with EXPECTED_PATH.open("r", encoding="utf-8") as handle:
        expected = json.load(handle)

    if headers != expected["source_columns"]:
        fail(f"Fixture headers differ from ground truth: {headers!r}")
    if len(rows) != expected["row_count"]:
        fail(f"Fixture row count is {len(rows)}, expected {expected['row_count']}")

    revenues = [int(row["Units"]) * int(row["Unit Price"]) for row in rows]
    if revenues != expected["revenue_by_row"]:
        fail(f"Fixture row revenues differ from ground truth: {revenues!r}")
    if sum(revenues) != expected["total_revenue"]:
        fail("Fixture total revenue differs from ground truth")
    if sum(int(row["Units"]) for row in rows) != expected["total_units"]:
        fail("Fixture total units differs from ground truth")

    for dimension, expected_key in (
        ("Region", "revenue_by_region"),
        ("Product", "revenue_by_product"),
    ):
        actual: dict[str, int] = {}
        for row, revenue in zip(rows, revenues):
            actual[row[dimension]] = actual.get(row[dimension], 0) + revenue
        if actual != expected[expected_key]:
            fail(f"Fixture {expected_key} differs: {actual!r}")

    return rows, expected


def find_header_row(ws: Any, required: list[str], max_rows: int = 10) -> tuple[int, dict[str, int]]:
    wanted = {norm(name): name for name in required}
    for row_index in range(1, min(ws.max_row, max_rows) + 1):
        found: dict[str, int] = {}
        for column_index in range(1, ws.max_column + 1):
            key = norm(ws.cell(row_index, column_index).value)
            if key in wanted:
                found[wanted[key]] = column_index
        if len(found) == len(required):
            return row_index, found
    fail(f"{ws.title}: could not find required headers {required!r}")


def compact_formula(value: Any) -> str:
    if not isinstance(value, str) or not value.startswith("="):
        return ""
    return value.upper().replace("$", "").replace(" ", "").replace("'", "")


def validate_revenue_formula(formula: Any, units_cell: str, price_cell: str) -> bool:
    compact = compact_formula(formula)
    if not compact or "*" not in compact:
        return False
    left, right = units_cell.upper(), price_cell.upper()
    if left in compact and right in compact:
        return True
    structured = compact.replace(" ", "")
    return "[@UNITS]" in structured and (
        "[@[UNITPRICE]]" in structured or "[@UNITPRICE]" in structured
    )


def find_label(ws: Any, label: str, max_rows: int = 30, max_columns: int = 20) -> Any:
    wanted = norm(label)
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, max_rows),
        min_col=1,
        max_col=min(ws.max_column, max_columns),
    ):
        for cell in row:
            if norm(cell.value) == wanted:
                return cell
    fail(f"{ws.title}: required label {label!r} not found")


def chart_ref(source: Any) -> str:
    if source is None:
        return ""
    for kind in ("strRef", "numRef"):
        ref = getattr(source, kind, None)
        formula = getattr(ref, "f", None) if ref is not None else None
        if formula:
            return compact_formula("=" + formula)[1:]
    return ""


def validate_workbook(path: Path, rows: list[dict[str, str]], expected: dict[str, Any]) -> list[str]:
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        fail("openpyxl is required to inspect the returned .xlsx")

    if not path.is_file():
        fail(f"Returned workbook not found: {path}")
    if path.suffix.lower() != ".xlsx":
        fail(f"Returned file must use .xlsx: {path}")

    try:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:
        fail(f"Workbook cannot be opened as .xlsx: {exc}")

    required_sheets = expected["required_sheets"]
    missing = [name for name in required_sheets if name not in workbook.sheetnames]
    if missing:
        fail(f"Required sheets missing: {missing!r}; found {workbook.sheetnames!r}")
    if len(workbook.sheetnames) != len(required_sheets):
        fail(f"Workbook must contain exactly {required_sheets!r}; found {workbook.sheetnames!r}")

    raw = workbook["Raw Data"]
    summary = workbook["Summary"]
    raw_headers = expected["source_columns"] + [expected["derived_column"]]
    header_row, columns = find_header_row(raw, raw_headers)

    data_rows: list[int] = []
    for row_index in range(header_row + 1, raw.max_row + 1):
        if any(raw.cell(row_index, columns[name]).value is not None for name in expected["source_columns"]):
            data_rows.append(row_index)
    if len(data_rows) != expected["row_count"]:
        fail(f"Raw Data preserves {len(data_rows)} source rows, expected {expected['row_count']}")

    for position, (row_index, source) in enumerate(zip(data_rows, rows), start=1):
        actual = {
            "Date": iso_date(raw.cell(row_index, columns["Date"]).value),
            "Region": str(raw.cell(row_index, columns["Region"]).value).strip(),
            "Product": str(raw.cell(row_index, columns["Product"]).value).strip(),
            "Units": str(int(number(raw.cell(row_index, columns["Units"]).value, f"row {position} Units"))),
            "Unit Price": str(int(number(raw.cell(row_index, columns["Unit Price"]).value, f"row {position} Unit Price"))),
        }
        if actual != source:
            fail(f"Raw Data row {position} differs: {actual!r} != {source!r}")

        units_cell = raw.cell(row_index, columns["Units"]).coordinate
        price_cell = raw.cell(row_index, columns["Unit Price"]).coordinate
        revenue_formula = raw.cell(row_index, columns["Revenue"]).value
        if not validate_revenue_formula(revenue_formula, units_cell, price_cell):
            fail(f"Raw Data row {position} Revenue is not a Units × Unit Price formula: {revenue_formula!r}")

    revenue_col = get_column_letter(columns["Revenue"])
    region_col = get_column_letter(columns["Region"])
    first_data_row, last_data_row = data_rows[0], data_rows[-1]
    raw_revenue_range = f"{revenue_col}{first_data_row}:{revenue_col}{last_data_row}".upper()
    raw_region_range = f"{region_col}{first_data_row}:{region_col}{last_data_row}".upper()

    total_label = find_label(summary, "Total Revenue")
    total_formula = summary.cell(total_label.row, total_label.column + 1).value
    compact_total = compact_formula(total_formula)
    if not compact_total or "SUM(" not in compact_total or "RAWDATA!" not in compact_total:
        fail(f"Summary Total Revenue must be a SUM formula referencing Raw Data: {total_formula!r}")
    if raw_revenue_range not in compact_total:
        fail(f"Summary Total Revenue formula must cover {raw_revenue_range}: {total_formula!r}")

    region_header = find_label(summary, "Region")
    revenue_header_col = None
    for column_index in range(1, summary.max_column + 1):
        value_norm = norm(summary.cell(region_header.row, column_index).value)
        if column_index != region_header.column and "revenue" in value_norm:
            revenue_header_col = column_index
            break
    if revenue_header_col is None:
        fail("Summary region table has no Revenue header")

    summary_rows: dict[str, int] = {}
    for row_index in range(region_header.row + 1, min(summary.max_row, region_header.row + 10) + 1):
        region_name = str(summary.cell(row_index, region_header.column).value or "").strip()
        if region_name in expected["revenue_by_region"]:
            summary_rows[region_name] = row_index
            formula = summary.cell(row_index, revenue_header_col).value
            compact = compact_formula(formula)
            criteria_cell = summary.cell(row_index, region_header.column).coordinate.upper()
            if not compact or not ("SUMIF(" in compact or "SUMIFS(" in compact):
                fail(f"{region_name} summary must use SUMIF/SUMIFS: {formula!r}")
            for required_reference in ("RAWDATA!", raw_region_range, raw_revenue_range, criteria_cell):
                if required_reference not in compact:
                    fail(f"{region_name} formula lacks {required_reference}: {formula!r}")
    if set(summary_rows) != set(expected["revenue_by_region"]):
        fail(f"Summary region rows differ: found {sorted(summary_rows)!r}")

    ordered_rows = [summary_rows[name] for name in expected["chart_contract"]["categories"]]
    if ordered_rows != list(range(min(ordered_rows), max(ordered_rows) + 1)):
        fail("Summary region rows must be contiguous and ordered North, Central, South")

    charts = list(getattr(summary, "_charts", []))
    if not charts:
        fail("Summary contains no chart")
    start_row, end_row = ordered_rows[0], ordered_rows[-1]
    category_range = (
        f"SUMMARY!{get_column_letter(region_header.column)}{start_row}:"
        f"{get_column_letter(region_header.column)}{end_row}"
    ).upper()
    value_range = (
        f"SUMMARY!{get_column_letter(revenue_header_col)}{start_row}:"
        f"{get_column_letter(revenue_header_col)}{end_row}"
    ).upper()
    matching_chart = False
    for chart in charts:
        if "bar" not in chart.__class__.__name__.lower():
            continue
        for series in chart.series:
            categories = chart_ref(getattr(series, "cat", None))
            values = chart_ref(getattr(series, "val", None))
            if category_range in categories and value_range in values:
                matching_chart = True
                break
    if not matching_chart:
        fail(
            "No column/bar chart uses the Summary region table ranges "
            f"{category_range} and {value_range}"
        )

    return [
        f"Workbook opens: {path}",
        "Raw Data: 12 source rows preserved in order",
        "Revenue: 12 real row formulas follow Units × Unit Price",
        f"Ground truth: total revenue {expected['total_revenue']}",
        "Summary: Total Revenue and region SUMIF/SUMIFS formulas reference Raw Data",
        "Chart: column/bar series references the Summary region table",
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--check-package",
        action="store_true",
        help="validate only the committed CSV/ground truth; do not claim provider execution",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"returned workbook path (default: {DEFAULT_WORKBOOK})",
    )
    args = parser.parse_args()

    try:
        rows, expected = load_package()
        print("PASS package: deterministic fixture matches ground truth")
        print(
            f"rows={expected['row_count']} units={expected['total_units']} "
            f"total_revenue={expected['total_revenue']}"
        )
        if args.check_package:
            print("NOT_EXECUTED: returned Claude workbook was intentionally not checked")
            return 0
        for message in validate_workbook(args.workbook.resolve(), rows, expected):
            print(f"PASS {message}")
        print("PASS technical workbook contract")
        print("AUTHORITY: evidence only; no automatic recommend or human approval")
        return 0
    except ContractError as exc:
        print(f"BLOCKED {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
